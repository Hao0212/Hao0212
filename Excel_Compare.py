from csv import excel
import re
import pandas as pd
import numpy as np
import math
import os 
import openpyxl 
from openpyxl.styles import PatternFill
from openpyxl.utils.cell import get_column_letter

def compare_folder(list1,list2):
    if len(list1) != len(list2):
        print("Number of files in the two folder are not equal")
        os._exit(0)
    
    list1_sorted = sorted(list1)
    list2_sorted = sorted(list2)
    
    for i in range(0,len(list1_sorted)):
        if list1_sorted[i] != list2_sorted[i]:
            print(list1_sorted[i] + " is different from " + list2_sorted[i])
            os._exit(0)


def compare_file(file1,file2):
    df1 = pd.ExcelFile(file1)
    df2 = pd.ExcelFile(file2)
    df1_sheet_name = sorted(df1.sheet_names)
    df2_sheet_name = sorted(df2.sheet_names)
    df1_sheet_number = len(df1_sheet_name)
    df2_sheet_number = len(df2_sheet_name)
    
    if(df1_sheet_number != df2_sheet_number):
        print(file1 + " has different sheet numbers !")
    else:
        for i in range(0,df1_sheet_number):
            if df1_sheet_name[i] != df2_sheet_name[i]:
                print("Sheet " + df1_sheet_name[i] + " is different from Sheet" + df2_sheet_name[i])
                break
            else:
                df1_sheet = pd.read_excel(file1, sheet_name = df1_sheet_name[i], header = None)
                df2_sheet = pd.read_excel(file2, sheet_name = df2_sheet_name[i], header = None)
                compare_sheet(df1_sheet,df2_sheet,df1_sheet_name[i],file1)
        

def compare_sheet(df1,df2,sheet_name,file1):
    if df1.shape[0] != df2.shape[0]:
        print(sheet_name,"Row number is different")
        return
    if df1.shape[1] != df2.shape[1]:
        print(sheet_name,"Column number is different")
        return
    # Comparison of two dataframe
    result = df1[df1 != df2].isna()
    index, column = np.where(result == False)
    # Change column's number to alphabet
    for i in range(0,len(index)):
            print(sheet_name, "is different on cell (",get_column_letter(column[i] + 1) + str(index[i] + 1),")")
            # Fill the different cell with red color
            excel_operation(file1,sheet_name,index,column)


def excel_operation(file1,sheet_name,index,column):
    fill_cell = PatternFill(patternType='solid',fgColor='FC2C03')
    file = openpyxl.load_workbook(file1)
    sheet = file[sheet_name]
    for i in range(0,len(index)):
        cell = get_column_letter(column[i] +1) + str(index[i]+1)
        sheet[cell].fill = fill_cell
    file.save(file1)


if  __name__ == '__main__':
    # Folder input
    print("Start Comparing Process\n")
    folder1_path = input("Please enter the first folder: ")
    folder1_excel_name = os.listdir(folder1_path)
    folder1_excel_name = sorted(folder1_excel_name)
    print("Files in the first folder:")
    print(folder1_excel_name)
    folder2_path = input("Please enter the second folder:  ")
    folder2_excel_name = os.listdir(folder2_path)
    folder2_excel_name = sorted(folder1_excel_name)

    # Compare the folder 
    compare_folder(folder1_excel_name, folder2_excel_name)
    print("Start comparing files in the folder\n")
    print("Files in the folder are the same\n")
    
    # Compare files
    print("Start comparing files: \n")
    for i in range(0,len(folder1_excel_name)):
        file_1 = folder1_path +"\\"+ folder1_excel_name[i]
        file_2 = folder2_path + "\\"+ folder2_excel_name[i]
        print()
        print("Start comparing", folder1_excel_name[i])
        compare_file(file_1,file_2)
      
